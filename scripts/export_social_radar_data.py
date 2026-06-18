from __future__ import annotations

import json
import os
from datetime import date, datetime
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "社媒热点题材雷达_v1.xlsx"
OUTPUT_PATH = ROOT / "dashboard" / "data" / "radar.json"
SNAPSHOT_DIR = ROOT / "dashboard" / "data" / "snapshots"
SNAPSHOT_INDEX_PATH = SNAPSHOT_DIR / "index.json"
YOUTUBE_REGIONS = {"US": "US", "UK": "GB", "CA": "CA", "AU": "AU"}
YOUTUBE_COMMENT_QUERIES = [
    "ReelShort short drama",
    "DramaBox short drama",
    "microdrama romance",
]
YOUTUBE_COMMENT_PAIN_PATTERNS = [
    {
        "pain_point": "YouTube观众抱怨短剧女主太软弱",
        "frequent_expression": "weak female lead / stop forgiving him / she should leave",
        "keywords": ["weak female", "weak fl", "forgive him", "forgiving him", "leave him", "stand up for herself"],
        "emotion": "anger, fatigue",
        "mapped_drama_angle": "不原谅爽剧 / 强女主反杀 / 离婚后逆袭",
        "recommended_handling": "前3集明确女主有主动目标和底线，减少反复原谅，把爽点放在证据、事业和新关系兑现。",
        "risk_notes": "避免只做羞辱前任，需让女主获得实际收益。",
    },
    {
        "pain_point": "YouTube观众对出轨复仇仍有强情绪",
        "frequent_expression": "cheater / revenge / expose him / karma",
        "keywords": ["cheater", "cheating", "revenge", "karma", "expose him", "expose her", "affair"],
        "emotion": "anger, anticipation",
        "mapped_drama_angle": "情感背叛 / 证据反杀 / 公开打脸",
        "recommended_handling": "保留出轨证据、公开打脸和财产反转，但增加职业身份或家庭秘密做差异化。",
        "risk_notes": "同质化高，不能只依赖背叛设定。",
    },
    {
        "pain_point": "YouTube观众讨厌付费后剧情拖水",
        "frequent_expression": "too many episodes / filler / where is the ending / paid but dragged",
        "keywords": ["too many episodes", "filler", "dragging", "dragged", "where is the ending", "full episode", "pay", "paid"],
        "emotion": "frustration",
        "mapped_drama_angle": "高密度反转 / 付费点后兑现 / 短集节奏优化",
        "recommended_handling": "每个付费段落必须兑现一个秘密、打脸或关系进展，避免用误会和重复冲突拖时长。",
        "risk_notes": "这是产品与节奏痛点，不直接等同于新题材。",
    },
    {
        "pain_point": "YouTube观众想看更有新意的狼人/命定伴侣",
        "frequent_expression": "rejected mate again / alpha / mate bond / werewolf story",
        "keywords": ["rejected mate", "alpha", "mate bond", "werewolf", "luna", "wolf"],
        "emotion": "fandom fatigue, curiosity",
        "mapped_drama_angle": "狼人 rejected mate / 世界观反转 / 女性成长",
        "recommended_handling": "保留命定拒绝爽点，但改成女主主动拒绝、规则反转或非alpha关系，降低同质化。",
        "risk_notes": "供给拥挤，必须有世界观或人物关系差异化。",
    },
    {
        "pain_point": "YouTube观众关注单亲/离婚后的重新开始",
        "frequent_expression": "single mom / divorce / kids / start over",
        "keywords": ["single mom", "single mother", "divorce", "divorced", "kids", "child", "children", "start over"],
        "emotion": "hope, empathy",
        "mapped_drama_angle": "单亲妈妈逆袭 / 离婚重启 / 家庭伦理",
        "recommended_handling": "把亲情、经济压力、前任纠缠和新身份揭示放在一起，前3集给明确生存压力和翻盘目标。",
        "risk_notes": "避免把女性成长只写成被拯救。",
    },
    {
        "pain_point": "YouTube观众对黑帮/高压关系既上头又担心有毒",
        "frequent_expression": "mafia romance / toxic / red flag / obsessed",
        "keywords": ["mafia", "toxic", "red flag", "possessive", "obsessed", "dangerous"],
        "emotion": "desire, concern",
        "mapped_drama_angle": "禁忌恋 / 危险保护欲 / 权力边界",
        "recommended_handling": "把危险感转成外部威胁和保护关系，弱化胁迫、囚禁和美化暴力。",
        "risk_notes": "品牌风险较高，只做弱信号追踪。",
    },
]


def clean(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def normalize_for_match(value):
    return str(value or "").lower()


def infer_audience_segment(item):
    text = normalize_for_match(" ".join(str(value) for value in item.values() if value))
    male_terms = [
        "男频", "male", "战神", "龙王", "赘婿", "son-in-law", "son in law",
        "final boss", "tech genius", "underdog", "sports", "hockey", "golf",
        "dragon slayer", "机甲", "西幻", "怪谈", "诡异", "王者", "末世", "竞技",
    ]
    female_terms = [
        "女频", "female", "女性", "romance", "billionaire", "heiress", "wife",
        "bride", "mom", "single mom", "alpha", "werewolf", "mafia", "forbidden",
        "contract", "cheating", "toxic relationship", "booktok", "追妻", "萌宝",
        "家庭伦理", "霸总", "情感", "禁忌恋",
    ]
    male_hit = any(term in text for term in male_terms)
    female_hit = any(term in text for term in female_terms)
    if male_hit and female_hit:
        return "混合/待验证"
    if male_hit:
        return "男频"
    if female_hit:
        return "女频"
    return "泛向/待验证"


def with_audience_segments(items):
    enriched = []
    for item in items or []:
        row = dict(item)
        row.setdefault("audience_segment", infer_audience_segment(row))
        enriched.append(row)
    return enriched


def sheet_records(workbook, sheet_name: str, header_row: int, start_row: int, max_blank_rows: int = 3):
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[header_row]]
    records = []
    blank_rows = 0

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if not any(cell is not None for cell in row):
            blank_rows += 1
            if blank_rows >= max_blank_rows:
                break
            continue

        blank_rows = 0
        record = {
            str(headers[idx]): clean(value)
            for idx, value in enumerate(row[: len(headers)])
            if headers[idx] is not None
        }
        records.append(record)

    return records


def risk_penalty(risk_notes):
    risk = str(risk_notes or "").lower()
    if "high" in risk:
        return 10
    if "medium" in risk:
        return 5
    return 0


def opportunity_score(cluster, angle):
    try:
        score = (
            0.35 * float(cluster.get("heat_score") or 0)
            + 0.25 * float(cluster.get("sentiment_score") or 0)
            + 0.20 * float(cluster.get("novelty_score") or 0)
            + 0.10 * float(cluster.get("sustainability_score") or 0)
            - risk_penalty(angle.get("risk_notes"))
        )
    except (TypeError, ValueError):
        return None
    return round(score)


def priority(score):
    if score is None:
        return None
    if score >= 70:
        return "High"
    if score >= 58:
        return "Watch"
    return "Risk/Low"


def hydrate_watchlist(watchlist, clusters, angles):
    clusters_by_name = {item.get("cluster_name"): item for item in clusters if item.get("cluster_name")}
    angles_by_name = {item.get("topic_cluster"): item for item in angles if item.get("topic_cluster")}
    hydrated = []

    for item in watchlist:
        topic = item.get("topic_cluster")
        cluster = clusters_by_name.get(topic, {})
        angle = angles_by_name.get(topic, {})
        score = item.get("opportunity_score")
        if score in (None, ""):
            score = opportunity_score(cluster, angle)

        hydrated.append(
            {
                **item,
                "priority": item.get("priority") or priority(score),
                "short_drama_genre": item.get("short_drama_genre") or angle.get("short_drama_genre"),
                "opportunity_score": score,
                "platforms_seen": item.get("platforms_seen") or cluster.get("platforms_seen"),
            }
        )

    return hydrated


BACKFILL_TOPIC_CLUSTERS = [
    {
        "cluster_name": "AI humanoid drama export",
        "trigger_event": "36氪和 DataEye 同时出现 AI短剧出海、AI仿真人剧、AI漫剧榜单信号。",
        "audience_pain_point": "想看更低成本但情绪密度接近真人短剧的新形态，尤其是逆袭、地域爽剧和高概念设定。",
        "dominant_emotion": "新奇 / 爽感 / 观望",
        "related_keywords": "AI humanoid drama; AI short drama export; ToonScroll; PineDrama; AI仿真人剧",
        "platforms_seen": "36Kr,DataEye",
        "heat_score": 72,
        "sentiment_score": 64,
        "novelty_score": 82,
        "sustainability_score": 68,
    },
    {
        "cluster_name": "Platform licensed microdrama",
        "trigger_event": "Peacock 引入 ReelShort 微短剧，海外主流流媒体开始测试竖屏短剧。",
        "audience_pain_point": "主流平台用户可能被短时长、高反转、强情绪内容吸引，但需要更精致的包装和更清晰类型。",
        "dominant_emotion": "好奇 / 试水 / 产业关注",
        "related_keywords": "Peacock ReelShort; microdrama licensing; Hollywood vertical drama; Bravo micro drama",
        "platforms_seen": "Business Insider,ReelShort,Peacock",
        "heat_score": 68,
        "sentiment_score": 58,
        "novelty_score": 78,
        "sustainability_score": 66,
    },
    {
        "cluster_name": "Destination inheritance romance",
        "trigger_event": "Hallmark 夏季 romance 与海外流媒体持续推出异国目的地、继承、契约条件类轻喜剧。",
        "audience_pain_point": "观众需要低门槛浪漫逃逸，同时保留继承、家族条件、旧爱重逢等明确冲突。",
        "dominant_emotion": "治愈 / 心动 / 逃离现实",
        "related_keywords": "destination romance; inheritance romance; Greek romance; contract marriage",
        "platforms_seen": "Hallmark,Decider,Netflix",
        "heat_score": 62,
        "sentiment_score": 60,
        "novelty_score": 58,
        "sustainability_score": 62,
    },
    {
        "cluster_name": "Underdog sports comeback",
        "trigger_event": "Netflix Top 10 与六月片单里出现体育逆袭、动画运动和经典拳击续作回流。",
        "audience_pain_point": "底层选手证明自己、被低估者翻盘和师徒成长仍有稳定爽感。",
        "dominant_emotion": "燃 / 逆袭 / 自我证明",
        "related_keywords": "sports comeback; underdog athlete; boxing revenge; training arc",
        "platforms_seen": "Netflix,Tom's Guide",
        "heat_score": 60,
        "sentiment_score": 62,
        "novelty_score": 54,
        "sustainability_score": 66,
    },
]


BACKFILL_WATCHLIST_SEEDS = [
    {
        "topic_cluster": "AI humanoid drama export",
        "priority": "Watch",
        "opportunity_score": 66,
        "short_drama_genre": "AI仿真人 / 出海试水 / 地域爽剧",
        "platforms_seen": "36Kr,DataEye",
        "watch_reason": "AI仿真人剧和AI短剧出海同时被行业媒体与榜单提及，适合先做题材与制作形态观察。",
        "recommended_action": "拆 AI 仿真人可接受度、地域爽剧 hook 和海外平台包装",
    },
    {
        "topic_cluster": "Platform licensed microdrama",
        "priority": "Watch",
        "opportunity_score": 61,
        "short_drama_genre": "平台授权 / Hollywood vertical drama",
        "platforms_seen": "Business Insider,Peacock,ReelShort",
        "watch_reason": "主流流媒体试水微短剧，说明竖屏短剧正在从独立App向平台内容库外溢。",
        "recommended_action": "跟踪 Peacock/ReelShort 授权内容的题材和用户反馈",
    },
    {
        "topic_cluster": "Destination inheritance romance",
        "priority": "Watch",
        "opportunity_score": 58,
        "short_drama_genre": "目的地恋爱 / 继承契约 / 轻喜剧",
        "platforms_seen": "Hallmark,Decider,Netflix",
        "watch_reason": "异国目的地、继承条件和契约婚恋组合低成本且容易做前3集冲突。",
        "recommended_action": "测试 inheritance + fake marriage + old flame 三类英文 hook",
    },
    {
        "topic_cluster": "Underdog sports comeback",
        "priority": "Risk/Low",
        "opportunity_score": 55,
        "short_drama_genre": "体育竞技 / 底层逆袭 / 训练爽感",
        "platforms_seen": "Netflix,Tom's Guide",
        "watch_reason": "体育逆袭情绪稳定，但海外短剧供给与制作成本需要再验证。",
        "recommended_action": "只做低成本拳击/格斗训练线观察，不碰大场面赛事",
    },
]


POTENTIAL_SIGNAL_SEEDS = [
    {
        "source_platform": "Times of India",
        "country_region": "IN / Global",
        "language": "EN",
        "keyword_or_hashtag": "Indian vertical OTT genre slate",
        "post_title_or_caption": "Rocket Reels adds eight original micro-dramas, including Lady Bond 007, Honey Trap, Marriage Bureau and action/crime/local comedy titles.",
        "url": "https://timesofindia.indiatimes.com/entertainment/hindi/bollywood/news/compelling-storytelling-can-thrive-in-shorter-formats-kranti-shanbhag/articleshow/131700683.cms",
        "views_or_rank": "2026-06-13 slate signal",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026-06 industry watch",
        "sentiment": "positive",
        "emotion_tags": "female spy, honey trap, marriage bureau, local comedy, action crime",
        "evidence_level": "B",
        "notes": "潜力题材：印度 vertical OTT 新片单显示女特工、诱捕悬疑、婚介轻喜、地方动作犯罪等非霸总/狼人方向也在扩张。",
    },
    {
        "source_platform": "Business Insider",
        "country_region": "US",
        "language": "EN",
        "keyword_or_hashtag": "AI absurd sci-fi microdrama",
        "post_title_or_caption": "TikTok Short Drama feed tests AI-created zombies, absurd comedy and sci-fi alongside crime and romance vertical series.",
        "url": "https://www.businessinsider.com/tiktok-testing-mini-drama-feed-ai-2026-3",
        "views_or_rank": "TikTok Short Drama feed signal",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 platform watch",
        "sentiment": "mixed",
        "emotion_tags": "AI, zombie, absurd comedy, sci-fi, crime",
        "evidence_level": "B",
        "notes": "潜力题材：AI荒诞科幻可测试低成本高概念钩子，如僵尸邻居、AI动物人、末日办公室和反常识生存规则。",
    },
    {
        "source_platform": "Business Insider",
        "country_region": "US",
        "language": "EN",
        "keyword_or_hashtag": "family drama interactive microdrama",
        "post_title_or_caption": "DramaBox is exploring family dramas and interactive content as it competes in the US microdrama race.",
        "url": "https://www.businessinsider.com/dramabox-seeks-new-funding-micro-drama-apps-gain-global-momentum-2026-1",
        "views_or_rank": "DramaBox strategy signal",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 platform watch",
        "sentiment": "positive",
        "emotion_tags": "family drama, interactive, mystery, inheritance",
        "evidence_level": "B",
        "notes": "潜力题材：家庭悬疑和互动分支适合做遗产争夺、失踪亲人、真假千金、评论投票选择下一集证据。",
    },
    {
        "source_platform": "Korean short-drama title watch",
        "country_region": "KR / Global",
        "language": "KO / EN",
        "keyword_or_hashtag": "palace assassin romance short drama",
        "post_title_or_caption": "Korean webtoon Gwang-an is being adapted into a short drama about a court maid assassin and a crown prince power-romance setup.",
        "url": "https://zh.wikipedia.org/wiki/%E7%8B%82%E7%9C%BC",
        "views_or_rank": "2026 webtoon short-drama adaptation",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 title watch",
        "sentiment": "positive",
        "emotion_tags": "webtoon IP, palace, assassin, romance, power struggle",
        "evidence_level": "C",
        "notes": "潜力题材：宫廷刺客恋爱可转成女刺客潜伏、暴君世子、身份暴露、权谋追杀，兼具女性向恋爱和动作悬疑。",
    },
    {
        "source_platform": "Business Insider",
        "country_region": "US / Brazil",
        "language": "EN",
        "keyword_or_hashtag": "free supernatural romance microdrama",
        "post_title_or_caption": "TikTok's PineDrama shows romance/supernatural microdramas; Love at First Bite and top trending shows signal vampire/werewolf appetite.",
        "url": "https://www.businessinsider.com/tiktok-launches-a-new-micro-drama-app-called-pinedrama-2026-1",
        "views_or_rank": "PineDrama platform signal",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 platform watch",
        "sentiment": "positive",
        "emotion_tags": "vampire, werewolf, romance, free viewing",
        "evidence_level": "B",
        "notes": "潜力题材：免费入口平台更适合测试高点击超自然恋爱，如吸血鬼初恋、狼人禁忌、命定伴侣反转。",
    },
    {
        "source_platform": "Business Insider",
        "country_region": "US",
        "language": "EN",
        "keyword_or_hashtag": "creator moral twist microdrama",
        "post_title_or_caption": "Fox and Dhar Mann Studios are producing 40 scripted vertical microdrama series for My Drama.",
        "url": "https://www.businessinsider.com/fox-partnering-with-dhar-mann-to-win-micro-drama-fans-2026-1",
        "views_or_rank": "studio/creator slate",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 studio watch",
        "sentiment": "positive",
        "emotion_tags": "creator-led, moral story, social lesson, reversal",
        "evidence_level": "B",
        "notes": "潜力题材：Dhar Mann式道德反转可转成短剧：被羞辱者翻盘、善恶报应、家庭/职场价值观冲突，每集一个明确教训+反转。",
    },
    {
        "source_platform": "Korean BL vertical drama list",
        "country_region": "KR / Global",
        "language": "KO / EN",
        "keyword_or_hashtag": "BL vertical short drama",
        "post_title_or_caption": "2026 Korean BL list includes vertical short dramas with 50+ episode formats.",
        "url": "https://zh.wikipedia.org/wiki/%E9%9F%A9%E5%9B%BDBL%E7%94%B5%E8%A7%86%E5%89%A7%E5%88%97%E8%A1%A8",
        "views_or_rank": "2026 title list",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 title watch",
        "sentiment": "positive",
        "emotion_tags": "BL, vertical, campus, youth romance",
        "evidence_level": "C",
        "notes": "潜力题材：BL竖屏短剧可观察校园暗恋、对立吸引、久别重逢、秘密关系等轻体量题材，适合海外小众但高粘性受众。",
    },
    {
        "source_platform": "Business Insider",
        "country_region": "US",
        "language": "EN",
        "keyword_or_hashtag": "brand microdrama commerce",
        "post_title_or_caption": "Brands are testing microdramas as shoppable entertainment.",
        "url": "https://www.wsj.com/cmo-today/body-scrub-concealer-and-shoe-trinkets-star-in-brands-microdramas-195af33c",
        "views_or_rank": "industry signal",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 industry watch",
        "sentiment": "mixed",
        "emotion_tags": "commerce, curiosity, brand risk",
        "evidence_level": "C",
        "notes": "潜力题材：品牌微短剧不一定映射到传统短剧类型，但可能变成广告投放和题材测试的新入口。",
    },
    {
        "source_platform": "arXiv / HCI",
        "country_region": "Global",
        "language": "EN",
        "keyword_or_hashtag": "viewer feedback driven writing",
        "post_title_or_caption": "Audience feedback-driven content creation in micro-drama production.",
        "url": "https://arxiv.org/abs/2602.14045",
        "views_or_rank": "research signal",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 research watch",
        "sentiment": "neutral",
        "emotion_tags": "production method, comments, iteration",
        "evidence_level": "C",
        "notes": "潜力题材：评论驱动创作更像生产方法论，可用于决定下一轮题材A/B测试。",
    },
    {
        "source_platform": "Axios",
        "country_region": "US",
        "language": "EN",
        "keyword_or_hashtag": "microdrama app funding",
        "post_title_or_caption": "Holywater raises funding for My Drama and microdrama expansion.",
        "url": "https://www.axios.com/2026/01/15/holywater-microdrama-app-funding",
        "views_or_rank": "funding signal",
        "likes": "",
        "comments": "",
        "shares": "",
        "trend_window": "2026 industry watch",
        "sentiment": "positive",
        "emotion_tags": "funding, platform competition, expansion",
        "evidence_level": "B",
        "notes": "潜力题材：平台融资本身不是题材，但说明海外微短剧供给侧仍在扩张，值得追踪平台偏好的类型。",
    },
]


COMMENT_PAIN_POINT_SEEDS = [
    {
        "pain_point": "不想再看软弱女主反复原谅",
        "frequent_expression": "Tired of weak female leads / why does she forgive him again",
        "source_platform": "Reddit / app reviews",
        "emotion": "anger, fatigue",
        "mapped_drama_angle": "不原谅爽剧 / 离婚后反杀 / 强女主复仇",
        "recommended_handling": "前3集明确女主不回头，男主追悔只能做阻力，爽点给事业和新关系。",
        "evidence_level": "B",
        "source_url": "https://www.reddit.com/search/?q=weak%20female%20lead%20microdrama",
        "risk_notes": "避免把女性成长简化成单纯羞辱前任，需补足主动目标。",
    },
    {
        "pain_point": "出轨复仇题材爽但同质化",
        "frequent_expression": "Cheating revenge is addictive but predictable",
        "source_platform": "TikTok / Reddit",
        "emotion": "anger, anticipation",
        "mapped_drama_angle": "情感背叛 / 复仇爽感 / 证据反杀",
        "recommended_handling": "保留出轨证据、公开打脸、财产反转，但增加职业身份或家庭秘密差异化。",
        "evidence_level": "B",
        "source_url": "https://www.reddit.com/search/?q=cheating%20revenge%20storytime",
        "risk_notes": "不能只堆羞辱，需要让女主有清晰收益和成长。",
    },
    {
        "pain_point": "熟龄女性二次恋爱供给少",
        "frequent_expression": "Need more stories about older women dating again",
        "source_platform": "Quora / Reddit / app reviews",
        "emotion": "hope, loneliness",
        "mapped_drama_angle": "熟龄二次恋爱 / 离婚重启 / 信任修复",
        "recommended_handling": "把恋爱线和子女、财务独立、前夫纠缠放在一起，避免只做年轻化甜宠。",
        "evidence_level": "C",
        "source_url": "https://www.quora.com/search?q=dating%20after%20divorce%20older%20women",
        "risk_notes": "需避免年龄羞辱和刻板化家庭压力。",
    },
    {
        "pain_point": "AI演员有违和感但剧情上头",
        "frequent_expression": "AI actors look creepy but the story is addictive",
        "source_platform": "Reddit / industry comments",
        "emotion": "curiosity, discomfort",
        "mapped_drama_angle": "AI仿真人剧可接受度 / 漫剧优先 / 高概念低成本测试",
        "recommended_handling": "优先用漫剧或半写实风格承接高概念题材，真人感AI只用于短镜头和强钩子测试。",
        "evidence_level": "C",
        "source_url": "https://www.reddit.com/search/?q=AI%20microdrama%20actors",
        "risk_notes": "不要夸大AI替代真人，页面只做内容形态观察。",
    },
    {
        "pain_point": "观众想参与剧情走向",
        "frequent_expression": "Let the comments decide what happens next",
        "source_platform": "arXiv / social comments",
        "emotion": "playful, participatory",
        "mapped_drama_angle": "评论共创 / 多结局恋爱 / 伪互动短剧",
        "recommended_handling": "把评论投票作为下一集钩子，不直接让用户决定核心价值观和合规风险情节。",
        "evidence_level": "C",
        "source_url": "https://arxiv.org/abs/2602.14045",
        "risk_notes": "评论反馈只能辅助A/B测试，不能替代编剧判断。",
    },
    {
        "pain_point": "黑帮/禁忌恋有吸引力但胁迫风险高",
        "frequent_expression": "Mafia romance is hot but too toxic",
        "source_platform": "BookTok / Reddit",
        "emotion": "desire, concern",
        "mapped_drama_angle": "禁忌恋 / 危险保护欲 / 权力边界",
        "recommended_handling": "把危险感转成外部威胁和保护关系，弱化胁迫、囚禁和美化暴力。",
        "evidence_level": "C",
        "source_url": "https://www.reddit.com/search/?q=mafia%20romance%20toxic",
        "risk_notes": "高品牌风险题材，只做弱信号追踪。",
    },
    {
        "pain_point": "用户讨厌付费后剧情拖水",
        "frequent_expression": "Too many filler episodes after I pay",
        "source_platform": "App Store / Google Play reviews",
        "emotion": "frustration",
        "mapped_drama_angle": "高密度反转 / 付费点后兑现 / 短集节奏优化",
        "recommended_handling": "每个付费段落必须兑现一个秘密、打脸或关系进展，避免只做误会延长。",
        "evidence_level": "B",
        "source_url": "https://play.google.com/store/search?q=ReelShort%20DramaBox&c=apps",
        "risk_notes": "这是产品/节奏痛点，不直接等同于新题材。",
    },
    {
        "pain_point": "想看更不套路的狼人命定伴侣",
        "frequent_expression": "Rejected mate again, but give me something new",
        "source_platform": "Reddit / TikTok comments",
        "emotion": "fandom fatigue, curiosity",
        "mapped_drama_angle": "狼人 rejected mate / 世界观新意 / 女性成长",
        "recommended_handling": "保留命定拒绝爽点，但改成女主主动拒绝、群体规则反转或非alpha关系。",
        "evidence_level": "C",
        "source_url": "https://www.reddit.com/search/?q=rejected%20mate%20werewolf%20romance",
        "risk_notes": "供给拥挤，必须有世界观或人物关系差异化。",
    },
]


STRATEGIC_FOCUS_SEEDS = [
    {
        "focus_name": "AI漫剧 / AI仿真人",
        "status": "重点关注",
        "audience_segment": "混合/待验证",
        "strategic_read": "AI漫剧适合承接高概念、奇幻、复仇、升级流和大世界观题材，可用较低制作成本测试真人短剧较难落地的设定。",
        "sample_signals": "DataEye短剧观察出现AI仿真人剧、2D漫剧和西幻/诡异题材上榜；36氪、Business Insider等媒体持续讨论AI短剧生产效率和AI演员争议。",
        "candidate_projects": "AI仿真人地域爽剧；IP改编AI漫剧；2D男频升级漫剧；诡异怪谈漫剧；评论改命/观众投票感短剧",
        "project_examples": [
            {
                "title": "AI仿真人地域爽剧",
                "platform": "题材发现",
                "date_window": "DataEye与行业稿均出现AI仿真人剧信号",
                "audience_segment": "混合/待验证",
                "topic_tag": "落魄千金、少夫人、隐婚曝光、地域身份反差",
                "evidence": "强身份反差+低成本场景，适合AI仿真人先测开头钩子",
            },
            {
                "title": "IP改编AI漫剧",
                "platform": "题材发现",
                "date_window": "36氪公开报道AI短剧出海平台动作",
                "audience_segment": "混合/待验证",
                "topic_tag": "网文IP、罗曼史IP、奇幻IP、竖屏漫剧",
                "evidence": "IP自带设定和人设，适合AI漫剧批量做题材A/B测试",
            },
            {
                "title": "2D男频升级漫剧",
                "platform": "题材发现",
                "date_window": "国内行业稿持续提到2D漫剧与播放增量",
                "audience_segment": "男频",
                "topic_tag": "西幻升级、机甲科幻、战力成长、怪谈副本",
                "evidence": "适合承接真人短剧成本较高的世界观和动作场面",
            },
            {
                "title": "诡异怪谈漫剧",
                "platform": "题材发现",
                "date_window": "行业稿出现诡异题材连续登顶表述",
                "audience_segment": "男频",
                "topic_tag": "怪谈规则、反常识设定、悬疑副本、强钩子开局",
                "evidence": "适合AI漫剧做高频视觉反转和低成本惊悚场景",
            },
            {
                "title": "评论改命/观众投票感短剧",
                "platform": "题材发现",
                "date_window": "2026年公开研究讨论观众反馈进入短剧创作",
                "audience_segment": "泛向/待验证",
                "topic_tag": "角色命运被评论改写、假互动投票、剧情分支、反向爽点兑现",
                "evidence": "适合包装成观众参与感强的竖屏故事，而不是只做制作方法",
            },
        ],
        "topic_directions": "AI仿真人地域爽剧; 2D男频升级; 东方奇幻/捉妖悬疑; 西幻升级; 诡异怪谈; 机甲科幻爽感",
        "validation_metrics": "优先记录项目名称、题材类型、榜单位置、播放/热度变化、投流素材增长、评论接受度和可复用英文hook。",
        "risk_notes": "AI演员替代和真人感质量仍有争议，适合作为趋势观察、样片验证和低成本题材测试，不宜直接包装成品牌主卖点。",
    },
    {
        "focus_name": "男频向短剧 / 漫剧",
        "status": "补强监控",
        "audience_segment": "男频",
        "strategic_read": "海外短剧供给仍以女频情感为主，男频可从高爽感、强目标、升级反馈和低成本动作/奇幻表达里寻找增量。",
        "sample_signals": "DataEye短剧/漫剧榜单、剧势分析、剧查查和公众号样本用于跟踪男频项目热度，重点看是否出现连续上榜、素材放量和评论共鸣。",
        "candidate_projects": "体育复出爽剧；失落继承人/家族复仇；赘婿反杀/强者回归；技术天才职场反杀；西幻升级/龙与王者；诡异怪谈/末世副本",
        "project_examples": [
            {
                "title": "体育复出爽剧",
                "platform": "题材发现",
                "date_window": "公开讨论出现高尔夫王者归来、Hockey Captain等体育男主样本",
                "audience_segment": "男频",
                "topic_tag": "退役王者、父女责任、学费压力、重返赛场、荣誉夺回",
                "evidence": "适合低成本训练线和强目标叙事，不依赖大场面赛事",
            },
            {
                "title": "失落继承人/家族复仇",
                "platform": "题材发现",
                "date_window": "公开报道出现 Lost Heir / Reckoning 类样本",
                "audience_segment": "男频",
                "topic_tag": "隐形继承人、家族压迫、身份揭晓、公开清算",
                "evidence": "适合前3集做羞辱-身份揭示-资源反杀",
            },
            {
                "title": "赘婿反杀/强者回归",
                "platform": "题材发现",
                "date_window": "国内行业稿和竞品样本均高频出现强身份反转",
                "audience_segment": "男频",
                "topic_tag": "被轻视丈夫、隐藏战神、技术天才、商业反杀",
                "evidence": "与海外观众熟悉的underdog comeback结构兼容",
            },
            {
                "title": "技术天才职场反杀",
                "platform": "题材发现",
                "date_window": "海外短剧已有 tech genius / fired genius 类样本",
                "audience_segment": "男频",
                "topic_tag": "被开除、核心代码、公司危机、前老板求回归",
                "evidence": "低成本办公室场景即可完成误判-危机-反杀闭环",
            },
            {
                "title": "西幻升级/龙与王者",
                "platform": "题材发现",
                "date_window": "DataEye出现西幻AI新剧、海外平台有龙/王者类题材样本",
                "audience_segment": "男频",
                "topic_tag": "龙、王者、魔法血统、被放逐后升级回归",
                "evidence": "适合漫剧化，真人拍摄成本高但视觉钩子强",
            },
            {
                "title": "诡异怪谈/末世副本",
                "platform": "题材发现",
                "date_window": "行业稿出现诡异题材高播放/登顶表述",
                "audience_segment": "男频",
                "topic_tag": "规则怪谈、末世副本、怪物机制、闯关生存",
                "evidence": "适合短集高频反转和低成本AI视觉测试",
            },
        ],
        "topic_directions": "战神/龙王归来; 都市赘婿逆袭; 西幻升级; 诡异怪谈; 末世机甲; 竞技训练爽剧; 退役王者回归",
        "validation_metrics": "优先记录竞品项目名、平台/榜单位置、受众向标签、付费爽点、前三集钩子、投流素材和海外评论反馈。",
        "risk_notes": "男频题材在海外短剧里需要验证文化转译和付费效率，先用榜单样本和投流素材判断，不直接替代女频主线。",
    },
]


def extend_clusters(clusters):
    existing = {item.get("cluster_name") for item in clusters if item.get("cluster_name")}
    extended = [dict(item) for item in clusters]
    for item in BACKFILL_TOPIC_CLUSTERS:
        if item["cluster_name"] not in existing:
            extended.append(dict(item))
    return extended


def extend_watchlist(watchlist):
    extended = [dict(item) for item in watchlist if item.get("topic_cluster")]
    existing = {item.get("topic_cluster") for item in extended}
    next_rank = max([int(item.get("rank") or 0) for item in extended] or [0]) + 1

    for item in BACKFILL_WATCHLIST_SEEDS:
        if len(extended) >= 10:
            break
        if item["topic_cluster"] in existing:
            continue
        extended.append({"rank": next_rank, **item})
        existing.add(item["topic_cluster"])
        next_rank += 1

    return extended


def build_signals(raw_signals, generated_at):
    signals = [dict(item) for item in raw_signals]
    existing_keys = {
        (item.get("keyword_or_hashtag"), item.get("source_platform"), item.get("country_region"))
        for item in signals
    }
    today = generated_at.date().isoformat()
    for item in POTENTIAL_SIGNAL_SEEDS:
        key = (item.get("keyword_or_hashtag"), item.get("source_platform"), item.get("country_region"))
        if key in existing_keys:
            continue
        signals.append({"date": today, "last_checked": today, **item})
    return signals


def build_comment_pain_points(youtube_points=None):
    points = []
    seen = set()
    for item in list(youtube_points or []) + [dict(seed) for seed in COMMENT_PAIN_POINT_SEEDS]:
        key = comment_pain_semantic_key(item)
        if not key or key in seen:
            continue
        seen.add(key)
        points.append(item)
    return points


def comment_pain_semantic_key(item):
    text = " ".join(
        str(item.get(field, "")).lower()
        for field in ("pain_point", "frequent_expression", "mapped_drama_angle")
    )
    semantic_rules = [
        ("weak-female-lead", ("weak female", "weak fl", "软弱女主", "反复原谅", "forgive him")),
        ("cheating-revenge", ("cheater", "cheating", "出轨", "复仇", "证据反杀")),
        ("paid-filler", ("filler", "drag", "付费", "拖水", "too many episodes")),
        ("werewolf-mate", ("werewolf", "rejected mate", "命定伴侣", "狼人", "alpha")),
        ("single-mom-divorce", ("single mom", "single mother", "单亲", "离婚后", "divorce")),
        ("mafia-toxic", ("mafia", "黑帮", "禁忌恋", "高压关系", "toxic", "red flag")),
    ]
    for key, tokens in semantic_rules:
        if any(token in text for token in tokens):
            return key
    return item.get("pain_point")


def build_strategic_focus():
    return [dict(item) for item in STRATEGIC_FOCUS_SEEDS]


def fetch_json(url, timeout=20):
    with urlopen(url, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def youtube_api_key():
    return os.environ.get("YOUTUBE_API_KEY", "").strip()


def fetch_youtube_trending_videos(generated_at, max_results_per_region=8):
    api_key = youtube_api_key()
    status = {
        "enabled": bool(api_key),
        "regions_requested": list(YOUTUBE_REGIONS.keys()),
        "regions_with_results": [],
        "error": "",
        "captured_at": generated_at.isoformat(timespec="seconds"),
    }
    if not api_key:
        status["error"] = "YOUTUBE_API_KEY not configured; skipped YouTube fetch."
        return [], status

    videos = []
    for display_region, api_region in YOUTUBE_REGIONS.items():
        params = {
            "part": "snippet,statistics",
            "chart": "mostPopular",
            "regionCode": api_region,
            "maxResults": max_results_per_region,
            "key": api_key,
        }
        url = f"https://www.googleapis.com/youtube/v3/videos?{urlencode(params)}"
        try:
            payload = fetch_json(url)
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as error:
            status["error"] = f"YouTube fetch failed for {display_region}: {error}"
            continue

        items = payload.get("items", [])
        if items:
            status["regions_with_results"].append(display_region)

        for idx, item in enumerate(items, start=1):
            snippet = item.get("snippet", {})
            statistics = item.get("statistics", {})
            video_id = item.get("id", "")
            videos.append(
                {
                    "date": generated_at.date().isoformat(),
                    "source_platform": "YouTube Trending",
                    "country_region": display_region,
                    "rank": idx,
                    "video_id": video_id,
                    "title": snippet.get("title", ""),
                    "channel_title": snippet.get("channelTitle", ""),
                    "published_at": snippet.get("publishedAt", ""),
                    "category_id": snippet.get("categoryId", ""),
                    "view_count": statistics.get("viewCount", ""),
                    "like_count": statistics.get("likeCount", ""),
                    "comment_count": statistics.get("commentCount", ""),
                    "url": f"https://www.youtube.com/watch?v={video_id}" if video_id else "",
                    "captured_at": generated_at.isoformat(timespec="seconds"),
                    "evidence_level": "A",
                }
            )

    return videos, status


def youtube_video_search_url(query):
    return f"https://www.youtube.com/results?{urlencode({'search_query': query})}"


def fetch_youtube_short_drama_videos(api_key, max_results_per_query=3):
    videos = []
    seen_ids = set()
    for query in YOUTUBE_COMMENT_QUERIES:
        params = {
            "part": "snippet",
            "type": "video",
            "q": query,
            "maxResults": max_results_per_query,
            "relevanceLanguage": "en",
            "safeSearch": "moderate",
            "key": api_key,
        }
        url = f"https://www.googleapis.com/youtube/v3/search?{urlencode(params)}"
        payload = fetch_json(url)
        for item in payload.get("items", []):
            video_id = (item.get("id") or {}).get("videoId")
            if not video_id or video_id in seen_ids:
                continue
            seen_ids.add(video_id)
            snippet = item.get("snippet", {})
            videos.append(
                {
                    "video_id": video_id,
                    "query": query,
                    "title": snippet.get("title", ""),
                    "channel_title": snippet.get("channelTitle", ""),
                    "url": f"https://www.youtube.com/watch?v={video_id}",
                }
            )
    return videos


def fetch_youtube_top_comments(api_key, video_id, max_results=12):
    params = {
        "part": "snippet",
        "videoId": video_id,
        "maxResults": max_results,
        "order": "relevance",
        "textFormat": "plainText",
        "key": api_key,
    }
    url = f"https://www.googleapis.com/youtube/v3/commentThreads?{urlencode(params)}"
    payload = fetch_json(url)
    comments = []
    for item in payload.get("items", []):
        snippet = (((item.get("snippet") or {}).get("topLevelComment") or {}).get("snippet") or {})
        text = snippet.get("textDisplay") or snippet.get("textOriginal") or ""
        if text:
            comments.append(text)
    return comments


def comment_matches(text, keywords):
    normalized = f" {str(text or '').lower()} "
    return any(keyword in normalized for keyword in keywords)


def fetch_youtube_comment_pain_points(generated_at):
    api_key = youtube_api_key()
    status = {
        "enabled": bool(api_key),
        "queries": list(YOUTUBE_COMMENT_QUERIES),
        "videos_checked": 0,
        "comments_scanned": 0,
        "matched_comments": 0,
        "error": "",
        "captured_at": generated_at.isoformat(timespec="seconds"),
    }
    if not api_key:
        status["error"] = "YOUTUBE_API_KEY not configured; skipped YouTube comment scan."
        return [], status

    aggregate = {
        pattern["pain_point"]: {
            "pattern": pattern,
            "matches": 0,
            "videos": set(),
            "queries": set(),
        }
        for pattern in YOUTUBE_COMMENT_PAIN_PATTERNS
    }

    try:
        videos = fetch_youtube_short_drama_videos(api_key)
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as error:
        status["error"] = f"YouTube search failed: {error}"
        return [], status

    for video in videos:
        status["videos_checked"] += 1
        try:
            comments = fetch_youtube_top_comments(api_key, video["video_id"])
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError):
            continue

        for comment in comments:
            status["comments_scanned"] += 1
            matched_any = False
            for item in aggregate.values():
                pattern = item["pattern"]
                if comment_matches(comment, pattern["keywords"]):
                    item["matches"] += 1
                    item["videos"].add(video["url"])
                    item["queries"].add(video["query"])
                    matched_any = True
            if matched_any:
                status["matched_comments"] += 1

    points = []
    for item in sorted(aggregate.values(), key=lambda value: value["matches"], reverse=True):
        if item["matches"] <= 0:
            continue
        pattern = item["pattern"]
        source_query = sorted(item["queries"])[0] if item["queries"] else YOUTUBE_COMMENT_QUERIES[0]
        evidence_level = "B" if item["matches"] >= 3 and len(item["videos"]) >= 2 else "C"
        points.append(
            {
                "pain_point": pattern["pain_point"],
                "frequent_expression": f"{pattern['frequent_expression']} · matched {item['matches']} public comments across {len(item['videos'])} videos",
                "source_platform": "YouTube public comments",
                "emotion": pattern["emotion"],
                "mapped_drama_angle": pattern["mapped_drama_angle"],
                "recommended_handling": pattern["recommended_handling"],
                "evidence_level": evidence_level,
                "source_url": youtube_video_search_url(source_query),
                "risk_notes": f"{pattern['risk_notes']} 已匿名聚合，不保留用户名或单条评论原文。",
            }
        )

    return points[:6], status


def youtube_videos_to_signals(videos):
    signals = []
    for item in videos:
        title = item.get("title", "")
        if not title:
            continue
        views = item.get("view_count") or "-"
        comments = item.get("comment_count") or "-"
        signals.append(
            {
                "date": item.get("date"),
                "source_platform": "YouTube Trending",
                "country_region": item.get("country_region"),
                "language": "EN",
                "keyword_or_hashtag": title[:120],
                "post_title_or_caption": title,
                "url": item.get("url"),
                "views_or_rank": f"rank {item.get('rank')} / views {views}",
                "likes": item.get("like_count"),
                "comments": comments,
                "shares": "",
                "trend_window": "daily YouTube mostPopular",
                "sentiment": "neutral",
                "emotion_tags": "video trend, public ranking",
                "evidence_level": item.get("evidence_level", "A"),
                "last_checked": item.get("captured_at", ""),
                "notes": f"YouTube mostPopular公开榜单：{item.get('channel_title', '')}",
            }
        )
    return signals


AI_ANIMATION_TOPIC_SEEDS = [
    {
        "topic": "《波斯复仇记》异域复仇AI短剧",
        "trend_signal": "36氪 2026年6月报道提到，AI短剧《波斯复仇记》成本约3000元，上线海外平台72小时GMV达50万美元，是近期可核验的AI短剧高回报案例线索。",
        "content_direction": "优先测试异域复仇、宫廷/贵族背叛、女性反杀、西幻权谋等视觉风格强、真人拍摄成本高的题材。",
        "audience_hook": "被流放的王妃回到宫廷，不再求爱，只带回能毁掉整个王室的秘密。",
        "source_platform": "36氪 / 镜像娱乐",
        "evidence_level": "B",
        "related_examples": "《波斯复仇记》, 异域复仇, 宫廷权谋, 女性反杀",
        "source_url": "https://www.36kr.com/p/3836864874560388",
        "risk_notes": "该案例来自行业报道，需继续验证平台、播放和付费口径；适合作为题材方向线索，不直接等同可复制爆款。",
    },
    {
        "topic": "ToonScroll网文IP漫剧出海",
        "trend_signal": "36氪 2026年6月报道提到，阅文推出海外漫剧平台 ToonScroll，计划年内上线超千部漫剧作品，显示网文IP正在进入大规模漫剧出海阶段。",
        "content_direction": "从网文IP里优先筛选西幻、复仇、强女主、萌宝、机甲、奇幻恋爱等短剧钩子明确的类型。",
        "audience_hook": "她以为自己穿进一本低分网文，却发现每改掉一个烂情节，现实里的仇人就会消失一个。",
        "source_platform": "36氪 / 镜像娱乐",
        "evidence_level": "B",
        "related_examples": "ToonScroll, 阅文IP, AI漫剧, 网文影像化",
        "source_url": "https://www.36kr.com/p/3836864874560388",
        "risk_notes": "平台计划需要后续用具体上线作品、播放表现和投流素材验证。",
    },
    {
        "topic": "FlareFlow / FlickReels AI仿真人剧本",
        "trend_signal": "36氪 2026年6月报道提到，中文在线 FlareFlow 转向AI短剧工厂化；FlickReels 与听花岛合作，重点扶持海外原创及AI仿真人剧本。",
        "content_direction": "先用AI仿真人测试契约婚姻、复仇、身份曝光、异域权谋、奇幻恋爱等强情绪题材，再决定真人化或漫剧化。",
        "audience_hook": "她签下AI替身婚约，却发现替身比真正的继承人更像她死去的爱人。",
        "source_platform": "36氪 / 镜像娱乐",
        "evidence_level": "B",
        "related_examples": "FlareFlow, FlickReels, AI仿真人剧本, 海外原创",
        "source_url": "https://www.36kr.com/p/3836864874560388",
        "risk_notes": "AI仿真人存在表演质感和伦理争议，前台包装应回到题材和故事，不宜只强调AI替代真人。",
    },
    {
        "topic": "AI荒诞科幻/僵尸短剧",
        "trend_signal": "TikTok Short Drama 测试入口中出现 AI 僵尸、荒诞动物人、科幻和犯罪等混合内容，显示平台在 romance 之外尝试更高概念题材。",
        "content_direction": "先做短剧化可理解的荒诞设定：僵尸邻居、AI动物人、末日办公室、会预言的AI同事、规则怪谈生存局。",
        "audience_hook": "第一集直接给反常识视觉钩子，第二集用生存规则或身份反转承接，而不是只靠奇观。",
        "source_platform": "Business Insider / TikTok Short Drama",
        "evidence_level": "B",
        "related_examples": "AI-generated zombies; absurd sci-fi; crime/romance feed mix",
        "source_url": "https://www.businessinsider.com/tiktok-testing-mini-drama-feed-ai-2026-3",
        "risk_notes": "荒诞题材容易只好笑不好追，需要绑定明确目标：逃生、复仇、恋爱或身份揭露。",
    },
    {
        "topic": "AI陪伴恋爱/人机情感替代",
        "trend_signal": "2026 韩剧 Love Phobia 把“用AI comfort替代人际连接”的设定放进浪漫喜剧和神秘秘密结构。",
        "content_direction": "短剧可转成女总裁沉迷AI男友、AI替身暴露真实秘密、真人小说家/程序员打破情感依赖。",
        "audience_hook": "她以为AI最懂自己，直到AI开始说出只有前任或死去亲人才知道的秘密。",
        "source_platform": "U+ Mobile TV / Lifetime",
        "evidence_level": "C",
        "related_examples": "Love Phobia",
        "source_url": "https://en.wikipedia.org/wiki/Love_Phobia_%28TV_series%29",
        "risk_notes": "要避免只做AI陪伴产品宣传，核心冲突应是亲密关系、秘密和自我修复。",
    },
    {
        "topic": "一句话生成短剧/个性化AI漫剧",
        "trend_signal": "2026 年新论文提出 One Sentence, One Drama，用多智能体流程从一句话生成短剧，并强调节奏、空间一致性和多阶段审校。",
        "content_direction": "把用户一句话偏好转成前3集钩子、角色关系、反转节点和视觉分镜，适合做AI漫剧题材A/B测试。",
        "audience_hook": "观众输入一个设定，系统生成专属爽剧：比如“被未婚夫背叛的女机甲师重回决赛”。",
        "source_platform": "arXiv / short-drama generation",
        "evidence_level": "C",
        "related_examples": "One Sentence, One Drama; Short-Drama-Bench",
        "source_url": "https://arxiv.org/abs/2605.22144",
        "risk_notes": "这是生产形态信号，需用真实投放和完播验证题材，不应只展示技术概念。",
    },
    {
        "topic": "创作者道德反转AI漫剧",
        "trend_signal": "Fox 与 Dhar Mann Studios 合作制作 40 部竖屏短剧，说明“道德教训+强反转”的 creator 叙事正在进入 microdrama。",
        "content_direction": "适合AI漫剧先测低成本社会寓言：拜金亲友羞辱、穷人身份反转、校园霸凌报应、职场恶老板翻车。",
        "audience_hook": "每集给一个明确恶行和即时反转，让观众获得报应爽感和评论争议点。",
        "source_platform": "Business Insider / Fox / Dhar Mann",
        "evidence_level": "B",
        "related_examples": "Dhar Mann Studios x My Drama",
        "source_url": "https://www.businessinsider.com/fox-partnering-with-dhar-mann-to-win-micro-drama-fans-2026-1",
        "risk_notes": "容易说教，必须保留人物欲望和连续悬念，不能只做单集道德短片。",
    },
    {
        "topic": "AI仿真人演员替代争议",
        "trend_signal": "Business Insider 近期报道美国 microdrama 产业中 AI 演员、AI短剧创业公司和真人演员替代争议升温。",
        "content_direction": "先用AI仿真人测试低成本强设定，如契约婚姻、复仇、身份曝光和奇幻恋爱；真人感不足时转向漫剧或半写实风格。",
        "audience_hook": "AI低成本 + 真人感角色 + 强情绪反转，适合测试观众对AI演员的接受度边界。",
        "source_platform": "Business Insider / industry discussion",
        "evidence_level": "B",
        "related_examples": "TrueShort, StoReel, AI-generated microdramas",
        "source_url": "https://www.businessinsider.com/actors-losing-jobs-to-ai-hollywood-micro-drama-industry-2026-6",
        "risk_notes": "AI演员替代有劳动伦理和表演质感争议，页面只做内容形态观察，不建议直接作为品牌主推卖点。",
    },
    {
        "topic": "单句生成短剧/多智能体生产",
        "trend_signal": "2026 年新论文提出 One Sentence, One Drama，用多智能体流程从一句话生成短剧，并强调节奏、空间一致性和多阶段审校。",
        "content_direction": "把题材发现、前3集钩子、分镜、BGM和审校拆成可复用流水线，适合AI漫剧和低成本概念样片。",
        "audience_hook": "用更快的题材A/B测试找到强hook，再决定是否进入真人或漫剧制作。",
        "source_platform": "arXiv / short-drama generation",
        "evidence_level": "C",
        "related_examples": "One Sentence, One Drama; Short-Drama-Bench",
        "source_url": "https://arxiv.org/abs/2605.22144",
        "risk_notes": "属于生产方法信号，不等同于用户真实需求；需要和评论痛点、投放表现共同验证。",
    },
    {
        "topic": "罗曼史IP改编漫剧",
        "trend_signal": "出版/网文IP正在被改造成竖屏 AI animated microdrama，讨论集中在 romance IP 是否适合 AI 漫剧化。",
        "content_direction": "契约婚姻、二次机会、灰姑娘逆袭、家族秘密等经典 romance trope，做成 30-60 秒强钩子连载。",
        "audience_hook": "熟悉的言情 trope + 漫画感视觉 + 快节奏反转。",
        "source_platform": "Publishers Weekly / Reddit / Dashverse",
        "evidence_level": "A/B",
        "related_examples": "Harlequin x Dashverse, A Fairy-Tail Ending",
        "source_url": "https://www.publishersweekly.com/pw/print/20260406/100043-harlequin-announces-slew-of-ai-generated-microdramas.html",
        "risk_notes": "AI 改编会引发读者对作者、画师和廉价感的争议，需强调审美与人工把关。",
    },
    {
        "topic": "东方奇幻/捉妖悬疑漫剧",
        "trend_signal": "平台级 AI 微动画案例开始跑出古风、捉妖、悬疑和世界观题材。",
        "content_direction": "女主误入妖局、少年除妖司、前世诅咒、门派背叛，适合竖屏连载和视觉奇观。",
        "audience_hook": "低成本做高概念世界观，用怪谈、禁术、反转身份撑开前 3 集。",
        "source_platform": "iQIYI / PRNewswire",
        "evidence_level": "A",
        "related_examples": "Imperial Exorcist Guards",
        "source_url": "https://www.prnewswire.com/news-releases/iqiyi-leads-ai-storytelling-as-original-micro-animations-gain-strong-traction-302713157.html",
        "risk_notes": "海外版本需要降低文化门槛，优先保留情感冲突和清晰视觉符号。",
    },
    {
        "topic": "机甲科幻爽感漫剧",
        "trend_signal": "AI 漫剧适合放大真人短剧成本较高的机甲、未来城市和超能力场面。",
        "content_direction": "废柴机师逆袭、AI 机甲觉醒、末世学院、战争遗孤复仇。",
        "audience_hook": "第一集给出机甲觉醒或战力爆发，后续用升级、背叛、队友情感维持追看。",
        "source_platform": "iQIYI / PRNewswire",
        "evidence_level": "A",
        "related_examples": "My Mecha is a Bit OP",
        "source_url": "https://www.prnewswire.com/news-releases/iqiyi-leads-ai-storytelling-as-original-micro-animations-gain-strong-traction-302713157.html",
        "risk_notes": "科幻设定容易解释成本高，必须用简单目标和强情绪降低理解门槛。",
    },
    {
        "topic": "漫画读者向奇幻恋爱",
        "trend_signal": "AI animated microdrama 的讨论常和 manga/webcomics 读者、fantasy romance 受众重合。",
        "content_direction": "狼人/吸血鬼、命定伴侣、魔法学院、异世界公主、禁忌恋等漫画感强类型。",
        "audience_hook": "高颜值角色设定 + 命定关系 + 每集一个情绪反转。",
        "source_platform": "Reddit / DramaGlance / Dashverse",
        "evidence_level": "B/C",
        "related_examples": "fantasy romance AI animation discussions",
        "source_url": "https://www.reddit.com/r/fantasyromance/comments/1s7v0o4/harlequin_to_coproduce_aigenerated_microdramas/",
        "risk_notes": "同质化风险高，需用新世界观或女性成长线区隔普通狼人/吸血鬼供给。",
    },
    {
        "topic": "互动分支/游戏IP漫剧",
        "trend_signal": "行业讨论把互动剧、游戏 IP 和 AI 动画短剧放在同一波内容形态里。",
        "content_direction": "多结局恋爱、乙女向选择、NPC觉醒、游戏反派重生、玩家进入剧情世界。",
        "audience_hook": "用选择感和强人设增加评论互动，适合做社媒投票和分支测试。",
        "source_platform": "industry discussion",
        "evidence_level": "C",
        "related_examples": "interactive drama, game-IP animation",
        "source_url": "https://min.news/en/news/53ea6d3712e03ebbbe5d8850ecbf5372.html",
        "risk_notes": "互动机制上线成本高，第一阶段可先做“伪互动”评论投票验证。",
    },
    {
        "topic": "微动效漫画/轻剧情漫剧",
        "trend_signal": "动画趋势讨论里，micro-animation 被认为适合把漫画、封面和插画变成短循环内容。",
        "content_direction": "治愈日常、暗恋瞬间、婚后小甜饼、复仇前夜、角色独白等轻剧情切片。",
        "audience_hook": "低成本高频测试视觉风格和角色吸引力，适合作为完整漫剧前的题材探针。",
        "source_platform": "CreativeBloq / animation trend blogs",
        "evidence_level": "B",
        "related_examples": "micro animation, motion comic loops",
        "source_url": "https://www.creativebloq.com/art/digital-art/digital-art-trends-2026-reveal-how-creatives-are-responding-to-ai-pressure",
        "risk_notes": "信息量太轻时不够短剧化，要补强人物目标和连续悬念。",
    },
    {
        "topic": "AI仿真人地域爽剧",
        "trend_signal": "DataEye 日榜出现 AI仿真人剧集中占位，行业媒体也在讨论 AI短剧出海和平台化。",
        "content_direction": "东北少夫人、落魄千金、乡镇逆袭、隐婚曝光等强情绪题材，用AI仿真人降低真人拍摄成本。",
        "audience_hook": "真人感角色 + 地域反差 + 每集身份反转，测试观众对AI演员的接受度。",
        "source_platform": "DataEye / 36Kr",
        "evidence_level": "B",
        "related_examples": "AI仿真人剧, ToonScroll, PineDrama",
        "source_url": "https://www.36kr.com/p/3836864874560388",
        "risk_notes": "AI真人感容易进入恐怖谷，需控制表演、口型和镜头时长。",
    },
    {
        "topic": "评论反馈驱动漫剧",
        "trend_signal": "2026 年研究开始关注 micro-drama 生产中如何使用评论、表情包和转发反馈反向影响剧情。",
        "content_direction": "先用低成本漫剧发布多版人物关系和前3集冲突，根据评论选择追妻、复仇、甜宠或悬疑方向。",
        "audience_hook": "让观众感觉参与了剧情走向，适合做社媒投票、评论点梗和多结局测试。",
        "source_platform": "arXiv / HCI",
        "evidence_level": "C",
        "related_examples": "Audience in the Loop",
        "source_url": "https://arxiv.org/abs/2602.14045",
        "risk_notes": "评论反馈只能辅助判断，不能替代基础叙事质量和合规审查。",
    },
]


TRADITIONAL_FILM_TV_TOPIC_SEEDS = [
    {
        "topic": "熟龄女性二次人生/小镇社区",
        "recent_signal": "Netflix《Sweet Magnolias》第5季于2026-06-11上线，并进入 Tom's Guide 对 Netflix 美国 Top 10 的推荐筛选，题材集中在熟龄女性友谊、家庭、婚礼、社区情感和小镇生活。",
        "short_drama_inspiration": "可转成30+女性二次人生短剧：离婚律师、单亲妈妈、婚礼策划师三位女性共享一个小镇秘密或家族遗产冲突。",
        "conversion_hook": "三位多年好友同时收到同一份遗嘱，才发现她们丈夫、前夫和初恋都卷入过同一场小镇丑闻。",
        "source_platform": "Netflix / Tom's Guide",
        "evidence_level": "B",
        "reference_titles": "Sweet Magnolias S5",
        "source_url": "https://www.tomsguide.com/entertainment/netflix/netflix-top-10-shows-heres-the-3-you-need-to-binge-watch-this-week-june-16-22",
        "risk_notes": "不要只做温情群像，短剧化需要在前3集给出遗产、婚礼危机、前任回归或小镇秘密等强冲突。",
    },
    {
        "topic": "CEO多重家庭/遗产爆雷",
        "recent_signal": "Netflix南非剧《The Polygamist》2026年6月上线并进入 Netflix Top 10 推荐讨论，Decider 单独评论其 CEO 多重关系、妻子/情人/秘密家庭、背叛复仇和葬礼倒叙结构。",
        "short_drama_inspiration": "可转成富豪男主死亡后，妻子、情人和秘密家庭同时出现在遗嘱宣读现场，围绕身份、财产、继承人和旧罪逐集爆雷。",
        "conversion_hook": "CEO葬礼当天，三个女人都以遗孀身份出现，而律师宣布真正继承人还没到场。",
        "source_platform": "Netflix / Decider",
        "evidence_level": "B",
        "reference_titles": "The Polygamist",
        "source_url": "https://decider.com/2026/06/12/the-polygamist-on-netflix-review/",
        "risk_notes": "多重关系题材要避免只做猎奇，核心应是多女性视角、家庭秘密和复仇/自救。",
    },
    {
        "topic": "宫廷刺客/权力恋爱短剧",
        "recent_signal": "韩国 webtoon《狂眼》被改编为短剧，设定是假扮宫女的刺客与性格乖戾世子之间的浪漫与权力斗争。",
        "short_drama_inspiration": "短剧可转成女刺客潜入王府/财阀家族，目标是刺杀或盗取证据，却被暴君继承人识破并被迫合作。",
        "conversion_hook": "女刺客第一晚潜入寝殿，却发现世子早已把她的通缉画像藏在枕下。",
        "source_platform": "Korean short-drama title watch",
        "evidence_level": "C",
        "reference_titles": "Gwang-an / 狂眼",
        "source_url": "https://zh.wikipedia.org/wiki/%E7%8B%82%E7%9C%BC",
        "risk_notes": "古装成本较高，海外短剧可转译成现代财阀宅邸、黑帮继承人或虚构小国王室。",
    },
    {
        "topic": "AI陪伴恋爱/情感替代",
        "recent_signal": "Love Phobia 以女 CEO 用 AI comfort 代替真实亲密关系为核心，和浪漫喜剧、神秘秘密结合。",
        "short_drama_inspiration": "短剧可把 AI 男友、情感替身、记忆数据和真人旧爱放在同一条线，做科技感情感悬疑。",
        "conversion_hook": "女主删除AI男友当天，AI却发来一段只有她失踪未婚夫才知道的语音。",
        "source_platform": "U+ Mobile TV / Lifetime",
        "evidence_level": "C",
        "reference_titles": "Love Phobia",
        "source_url": "https://en.wikipedia.org/wiki/Love_Phobia_%28TV_series%29",
        "risk_notes": "重点是情感悬疑，不是AI产品展示；需要处理隐私和依赖风险。",
    },
    {
        "topic": "BL奇幻命运重启",
        "recent_signal": "台湾 BL 奇幻剧 Wishing Upon the Shooting Stars 在 2026 年播出，并通过 Viki、GagaOOLala 等面向国际观众。",
        "short_drama_inspiration": "可转成海外 BL 短剧：流星许愿、平行时空、错过的暗恋重来一次、命运倒计时。",
        "conversion_hook": "男主许愿回到十年前，却发现每救一次暗恋对象，自己就会少活一天。",
        "source_platform": "LINE TV / Viki / GagaOOLala",
        "evidence_level": "C",
        "reference_titles": "Wishing Upon the Shooting Stars",
        "source_url": "https://en.wikipedia.org/wiki/Wishing_Upon_the_Shooting_Stars",
        "risk_notes": "BL题材受众粘性高但圈层化，适合小成本高情绪测试，不宜泛化为大众盘。",
    },
    {
        "topic": "神秘转学生/校园审判惊悚",
        "recent_signal": "日本版《Girl from Nowhere》在 2026 年上线，延续神秘少女进入校园、揭露学生和教师阴暗面的单元审判结构。",
        "short_drama_inspiration": "短剧可转成每 6-8 集一个校园/职场单元：神秘新生、实习生或清洁工进入封闭圈层，诱发霸凌、作弊、权色交易和秘密曝光。",
        "conversion_hook": "转学生第一天就知道全班最受欢迎女生的秘密，并把真相做成匿名投票。",
        "source_platform": "FOD / Fuji TV / Wikipedia",
        "evidence_level": "C",
        "reference_titles": "Transfer Student Nanno / Girl from Nowhere remake",
        "source_url": "https://zh.wikipedia.org/wiki/%E8%BD%89%E5%AD%B8%E4%BE%86%E7%9A%84%E5%A5%B3%E7%94%9F_%28%E6%97%A5%E6%9C%AC%E9%9B%BB%E8%A6%96%E5%8A%87%29",
        "risk_notes": "校园题材需规避未成年人伤害细节，重点放在审判机制、悬疑反转和成年化职场替代版本。",
    },
    {
        "topic": "本地社区女性群像/阶层喜剧",
        "recent_signal": "Guardian 近期全球观看观察提到科特迪瓦剧集 Les Nounous 以保姆群体、地方语言和社区生活形成国民级讨论。",
        "short_drama_inspiration": "可转成移民保姆、月嫂、清洁工、护理员等女性劳动者群像：她们掌握雇主秘密、互助反击、解决家庭危机。",
        "conversion_hook": "新来的保姆第一天发现豪门太太失踪，所有雇主都以为她只会沉默。",
        "source_platform": "The Guardian",
        "evidence_level": "B",
        "reference_titles": "Les Nounous",
        "source_url": "https://www.theguardian.com/culture/2026/may/22/from-chinese-microdramas-to-an-arctic-comedy-what-the-world-is-watching",
        "risk_notes": "要避免刻板化底层职业，重点做女性互助、阶层观察和秘密反转。",
    },
    {
        "topic": "成名塌房/依赖型BL重逢",
        "recent_signal": "2026 年 BL 剧《Love After Addiction》围绕演员事业崩塌、失眠依赖和旧同学重逢展开，说明名人危机+情感依赖仍有讨论度。",
        "short_drama_inspiration": "可转成海外BL/泛向情感短剧：塌房明星回到小镇，只能在旧友身边入睡，旧秘密和职业复出同步推进。",
        "conversion_hook": "过气演员直播崩溃后躲回老家，却发现唯一能让他入睡的人正是当年被他抛弃的同学。",
        "source_platform": "GagaOOLala / Wikipedia",
        "evidence_level": "C",
        "reference_titles": "Love After Addiction",
        "source_url": "https://en.wikipedia.org/wiki/Love_After_Addiction",
        "risk_notes": "涉及成瘾、失眠和心理依赖时要避免美化病理关系，可改成治愈互助和事业复出。",
    },
    {
        "topic": "巨制奇幻权谋/家族内战",
        "recent_signal": "海外长视频六月档继续押注高识别度奇幻IP和战争级大事件，适合观察“家族权力斗争+血缘背叛”的耐久需求。",
        "short_drama_inspiration": "把宏大战争压缩成家族继承、私生子身份、婚约联盟、背叛复仇等低成本关系冲突。",
        "conversion_hook": "第一集直接给出继承权被夺、婚盟背刺或血统秘密曝光。",
        "source_platform": "HBO Max / Tom's Guide",
        "evidence_level": "B",
        "reference_titles": "House of the Dragon S3",
        "source_url": "https://www.tomsguide.com/entertainment/streaming/20-new-shows-and-movies-to-watch-in-june-2026-the-bear-house-of-the-dragon-and-more",
        "risk_notes": "短剧不宜硬做大场面，优先转译成权谋关系、身份反转和高密度悬念。",
    },
    {
        "topic": "BookTok/女性向小说改编恋爱",
        "recent_signal": "海外平台持续推出文学/畅销书改编 romance，BookTok 受众和流媒体剧集形成互相导流。",
        "short_drama_inspiration": "二次机会、夏日旧爱、久别重逢、校园到成人的遗憾补偿，适合女性向短剧连续追更。",
        "conversion_hook": "多年后重逢时，女主发现男主一直保存当年的秘密证据或未寄出的信。",
        "source_platform": "Prime Video / AP / Tom's Guide",
        "evidence_level": "B",
        "reference_titles": "Every Year After",
        "source_url": "https://apnews.com/article/19b8e5b96c6860404d16c0cebb19c369",
        "risk_notes": "要避免平铺直叙，前 3 集必须把旧误会、现任阻碍和现实利益放上桌。",
    },
    {
        "topic": "职场恋爱/强势女主失控感",
        "recent_signal": "海外六月新片仍在推明星驱动的职场 rom-com，办公室权力差与情感失控是可短剧化的强冲突。",
        "short_drama_inspiration": "女 CEO、律师、空降高管、竞争对手等设定，转成强势女主与危险新人/旧爱之间的拉扯。",
        "conversion_hook": "女主刚宣布绝不办公室恋情，新来的法务却掌握她公司最大危机。",
        "source_platform": "Netflix / Tom's Guide",
        "evidence_level": "B",
        "reference_titles": "Office Romance",
        "source_url": "https://www.tomsguide.com/entertainment/streaming/20-new-shows-and-movies-to-watch-in-june-2026-the-bear-house-of-the-dragon-and-more",
        "risk_notes": "需处理好职场权力不对等，避免让爱情线显得越界或不适。",
    },
    {
        "topic": "家庭法庭/婚姻纠纷现实题材",
        "recent_signal": "国内长剧出现家庭法庭、婚姻纠纷、抚养权、养老与继承等社会议题型法治剧。",
        "short_drama_inspiration": "把一个案件压缩成一条高情绪主线：离婚争产、抚养权争夺、老人遗嘱、家暴取证、隐婚曝光。",
        "conversion_hook": "女主上庭争抚养权时，发现丈夫提交的证据来自她最信任的亲人。",
        "source_platform": "Tencent Video / iQIYI / IMDb",
        "evidence_level": "B",
        "reference_titles": "Hold a Court Now",
        "source_url": "https://www.imdb.com/title/tt30478333/plotsummary/",
        "risk_notes": "涉及家暴、未成年人和法律流程时要弱化猎奇，保留情绪但避免误导法律常识。",
    },
    {
        "topic": "古装女性凝视/强女主复仇",
        "recent_signal": "国内古偶爆款继续验证女性向古装、反差人设、婚恋权谋和女主主动性的组合。",
        "short_drama_inspiration": "将古装大剧的朝堂/门阀简化为婚约、替嫁、复仇、身份调换和双强博弈。",
        "conversion_hook": "替嫁当晚，女主发现新郎正是灭门旧案唯一幸存者。",
        "source_platform": "Tencent Video / iQIYI / Netflix / Tonboriday",
        "evidence_level": "B",
        "reference_titles": "Pursuit of Jade",
        "source_url": "https://www.tonboriday.com/2026/03/pursuit-of-jade-becomes-2026s-breakout.html",
        "risk_notes": "短剧化时要减少朝代设定解释，把爽点集中在身份、婚约和复仇反转。",
    },
    {
        "topic": "明星传记/名人代价",
        "recent_signal": "院线与流媒体持续消化明星传记片热度，观众对天才、家庭控制、名誉危机和遗产争议有稳定兴趣。",
        "short_drama_inspiration": "可转译为虚构娱乐圈：天才偶像、控制型家人、经纪公司压榨、旧丑闻反噬、舞台复出。",
        "conversion_hook": "过气女星复出演唱会前夜，收到一段足以毁掉她职业生涯的旧录像。",
        "source_platform": "Theatrical / PVOD / AP",
        "evidence_level": "B",
        "reference_titles": "Michael",
        "source_url": "https://apnews.com/article/19b8e5b96c6860404d16c0cebb19c369",
        "risk_notes": "必须虚构化，避免影射真实人物争议；重点放在名利场情绪而非真实八卦。",
    },
    {
        "topic": "目的地继承/契约浪漫",
        "recent_signal": "Hallmark 夏季 romance 继续押注异国目的地、继承条件、临时婚约和轻喜剧误会。",
        "short_drama_inspiration": "把目的地风景压缩成继承条款、假结婚、旧爱重逢和家族餐厅危机。",
        "conversion_hook": "女主到希腊继承餐厅，却被告知必须和陌生共同继承人结婚才能拿到产权。",
        "source_platform": "Hallmark / Decider",
        "evidence_level": "B",
        "reference_titles": "The Greek Aisle",
        "source_url": "https://decider.com/2026/06/06/the-greek-aisle-hallmark-cast-guide/",
        "risk_notes": "目的地元素不能只做风景，要在前3集放入明确利益冲突和情感阻碍。",
    },
    {
        "topic": "体育逆袭/训练爽剧",
        "recent_signal": "Netflix Top 10 和六月片单里体育动画、拳击经典和底层选手证明自己的内容仍有热度。",
        "short_drama_inspiration": "低成本转成拳馆、地下格斗、女主教练、伤病复出和家族债务。",
        "conversion_hook": "被禁赛的女拳手回到破旧拳馆，发现新学员正是当年害她退役的人。",
        "source_platform": "Netflix / Tom's Guide",
        "evidence_level": "B",
        "reference_titles": "Goat, Creed",
        "source_url": "https://www.tomsguide.com/entertainment/netflix/netflix-top-10-movies-heres-the-3-worth-watching-right-now-june-6-7",
        "risk_notes": "比赛场面成本较高，优先做训练、债务、师徒和复仇关系线。",
    },
]


INDUSTRY_MEDIA_OBSERVATION_SEEDS = [
    {
        "source_name": "Times of India",
        "article_date": "2026-06-13",
        "title": "Rocket Reels expands with eight new original micro-dramas",
        "summary": "印度 vertical OTT 平台 Rocket Reels 在一次性推出 10 部原创剧后继续扩张，新增 Lady Bond 007、Honey Trap、Marriage Bureau、Pathan The Killer 等 8 部原创 micro-drama。",
        "topic_signal": "印度竖屏短剧 / 女特工 / 诱捕悬疑 / 婚介轻喜 / 动作犯罪",
        "evidence_level": "B",
        "source_url": "https://timesofindia.indiatimes.com/entertainment/hindi/bollywood/news/compelling-storytelling-can-thrive-in-shorter-formats-kranti-shanbhag/articleshow/131700683.cms",
    },
    {
        "source_name": "Business Insider",
        "article_date": "2026-03-18",
        "title": "TikTok tests Short Drama feed with AI zombies and sci-fi",
        "summary": "TikTok 主站测试 Short Drama feed，内容覆盖 crime、romance、absurd comedy、sci-fi，并出现 AI 僵尸等高概念内容，可观察平台是否把短剧从 romance 推向更广类型。",
        "topic_signal": "AI荒诞科幻 / 僵尸短剧 / TikTok短剧入口",
        "evidence_level": "B",
        "source_url": "https://www.businessinsider.com/tiktok-testing-mini-drama-feed-ai-2026-3",
    },
    {
        "source_name": "Business Insider",
        "article_date": "2026-01-22",
        "title": "DramaBox explores family dramas and interactive content",
        "summary": "DramaBox 在美国竞争中被提到探索 family dramas 和 interactive content，可作为遗产、亲情悬疑、真假亲人和评论分支题材观察。",
        "topic_signal": "家庭悬疑 / 互动分支 / 遗产与亲情反转",
        "evidence_level": "B",
        "source_url": "https://www.businessinsider.com/dramabox-seeks-new-funding-micro-drama-apps-gain-global-momentum-2026-1",
    },
    {
        "source_name": "Business Insider",
        "article_date": "2026-01-08",
        "title": "Microdramas face reality check after US arrival",
        "summary": "报道指出 Fox、Disney 等玩家正在探索 romance 之外的 crime、animation、广告支持和产品植入模式，说明题材需要从霸总/狼人扩展到犯罪、动画和品牌剧情。",
        "topic_signal": "犯罪短剧 / 动画短剧 / 品类扩张",
        "evidence_level": "B",
        "source_url": "https://www.businessinsider.com/micro-dramas-face-reality-check-after-buzzy-us-arrival-2026-1",
    },
    {
        "source_name": "Korean short-drama title watch",
        "article_date": "2026-04-28",
        "title": "Gwang-an webtoon adapted into short drama",
        "summary": "韩国人气 webtoon《狂眼》改编短剧，设定为假扮宫女的刺客与世子之间的浪漫、权力和身份斗争。",
        "topic_signal": "宫廷刺客 / webtoon IP / 权力恋爱",
        "evidence_level": "C",
        "source_url": "https://zh.wikipedia.org/wiki/%E7%8B%82%E7%9C%BC",
    },
    {
        "source_name": "Korean drama title watch",
        "article_date": "2026-02-19",
        "title": "Love Phobia centers AI comfort and romantic mystery",
        "summary": "Love Phobia 将 AI comfort、人际连接缺失、浪漫喜剧和神秘秘密结合，可转化为AI陪伴恋爱/情感替代短剧方向。",
        "topic_signal": "AI陪伴恋爱 / 情感替代 / 科技悬疑",
        "evidence_level": "C",
        "source_url": "https://en.wikipedia.org/wiki/Love_Phobia_%28TV_series%29",
    },
    {
        "source_name": "Taiwan BL title watch",
        "article_date": "2026-06-04",
        "title": "Wishing Upon the Shooting Stars finishes run",
        "summary": "台湾 BL 奇幻剧通过 Viki、GagaOOLala 面向国际观众，流星许愿、平行命运和错过重来可作为 BL 短剧轻奇幻方向。",
        "topic_signal": "BL奇幻 / 命运重启 / 国际小众受众",
        "evidence_level": "C",
        "source_url": "https://en.wikipedia.org/wiki/Wishing_Upon_the_Shooting_Stars",
    },
    {
        "source_name": "Business Insider",
        "article_date": "2026-01-16",
        "title": "TikTok launches PineDrama in the US and Brazil",
        "summary": "PineDrama 以免费、无广告试水 microdrama，热门内容集中在 romance 和 supernatural，Love at First Bite 等吸血鬼/狼人恋爱题材具备短剧化参考价值。",
        "topic_signal": "免费平台测试 / 吸血鬼恋爱 / 超自然romance",
        "evidence_level": "B",
        "source_url": "https://www.businessinsider.com/tiktok-launches-a-new-micro-drama-app-called-pinedrama-2026-1",
    },
    {
        "source_name": "Business Insider",
        "article_date": "2026-01-26",
        "title": "Fox partners with Dhar Mann for microdrama slate",
        "summary": "Fox 与 Dhar Mann Studios 合作 40 部 scripted vertical microdrama，说明 creator-led 的道德反转、社会议题和强报应叙事正在被主流公司吸收。",
        "topic_signal": "道德反转 / creator-led短剧 / 报应爽感",
        "evidence_level": "B",
        "source_url": "https://www.businessinsider.com/fox-partnering-with-dhar-mann-to-win-micro-drama-fans-2026-1",
    },
    {
        "source_name": "arXiv",
        "article_date": "2026-05-21",
        "title": "One Sentence, One Drama",
        "summary": "论文提出用多智能体从一句话生成短剧，并强调短剧节奏、钩子、空间一致性和多阶段审校，可作为 AI 漫剧个性化生产方向。",
        "topic_signal": "一句话生成短剧 / 个性化AI漫剧 / 多智能体",
        "evidence_level": "C",
        "source_url": "https://arxiv.org/abs/2605.22144",
    },
    {
        "source_name": "The Guardian",
        "article_date": "2026-05-22",
        "title": "What the world is watching",
        "summary": "全球观看观察里提到中国 microdrama、阿根廷女性分手喜剧、加拿大北极社区喜剧和科特迪瓦保姆群像，说明本地文化和女性社区故事仍能形成讨论。",
        "topic_signal": "本地社区题材 / 女性劳动者群像 / 分手后自我重建",
        "evidence_level": "B",
        "source_url": "https://www.theguardian.com/culture/2026/may/22/from-chinese-microdramas-to-an-arctic-comedy-what-the-world-is-watching",
    },
    {
        "source_name": "FOD / Fuji TV",
        "article_date": "2026-02-20",
        "title": "Girl from Nowhere Japanese remake announced",
        "summary": "泰国《禁忌女孩》日本版改编以神秘转学生揭露校园黑暗面为核心，可转译成短剧里的校园/职场审判单元剧。",
        "topic_signal": "神秘转学生 / 校园审判 / 单元悬疑",
        "evidence_level": "C",
        "source_url": "https://zh.wikipedia.org/wiki/%E8%BD%89%E5%AD%B8%E4%BE%86%E7%9A%84%E5%A5%B3%E7%94%9F_%28%E6%97%A5%E6%9C%AC%E9%9B%BB%E8%A6%96%E5%8A%87%29",
    },
    {
        "source_name": "Korean BL title watch",
        "article_date": "2026-03-05",
        "title": "2026 Korean BL includes vertical short dramas",
        "summary": "2026 韩国 BL 列表出现 50 集以上竖屏短剧形态，校园暗恋、对立吸引和久别重逢可作为海外小众高粘性题材观察。",
        "topic_signal": "BL竖屏短剧 / 校园暗恋 / 久别重逢",
        "evidence_level": "C",
        "source_url": "https://zh.wikipedia.org/wiki/%E9%9F%A9%E5%9B%BDBL%E7%94%B5%E8%A7%86%E5%89%A7%E5%88%97%E8%A1%A8",
    },
    {
        "source_name": "Business Insider",
        "article_date": "2026-06-05",
        "title": "Actors booked starring roles, then were replaced by AI",
        "summary": "报道显示美国 microdrama 产业开始用 AI 生成角色和剧集压低成本，演员替代、表演可信度和观众接受度正在成为新争议。",
        "topic_signal": "AI演员替代 / 真人感短剧 / 生产成本与伦理风险",
        "evidence_level": "B",
        "source_url": "https://www.businessinsider.com/actors-losing-jobs-to-ai-hollywood-micro-drama-industry-2026-6",
    },
    {
        "source_name": "arXiv",
        "article_date": "2026-05-21",
        "title": "One Sentence, One Drama",
        "summary": "论文提出用多智能体系统把一句话扩展成短剧，重点解决短剧节奏、跨镜头空间一致性和多阶段审校，可作为AI短剧生产方法观察。",
        "topic_signal": "AI短剧生成 / 多智能体流程 / Short-Drama-Bench",
        "evidence_level": "C",
        "source_url": "https://arxiv.org/abs/2605.22144",
    },
    {
        "source_name": "36氪 / 骨朵网络影视",
        "article_date": "2026-06-04",
        "title": "ROI碾压，才是AI漫剧对真人剧的降维攻击",
        "summary": "文章讨论漫剧投流、播放和ROI表现，指出网文IP、低成本制作和快速周转正在推动漫剧成为短剧赛道的重要变量。",
        "topic_signal": "AI漫剧 / 网文IP影像化 / ROI与资金周转",
        "evidence_level": "B",
        "source_url": "https://www.36kr.com/p/3838601706440965",
    },
    {
        "source_name": "36氪 / 文娱价值官",
        "article_date": "2026-06-03",
        "title": "观众越来越挑食，出海短剧也要卷“细糠”了",
        "summary": "文章强调海外短剧从翻译搬运进入深度本土化阶段，北美用户对套路化内容耐受降低，题材需要更贴近当地文化语境。",
        "topic_signal": "出海短剧本土化 / 北美审美变化 / 精品化",
        "evidence_level": "B",
        "source_url": "https://www.36kr.com/p/3837211878947464",
    },
    {
        "source_name": "36氪 / 镜像娱乐",
        "article_date": "2026-06-03",
        "title": "AI短剧出海热：营收翻1200倍、订单暴涨5000%，阅文字节中文在线等公司纷纷入局",
        "summary": "文章提到《波斯复仇记》低成本AI短剧案例、阅文 ToonScroll 海外漫剧平台、字节 PineDrama/MiniShorts、TikTok minis、中文在线 FlareFlow 和 FlickReels/听花岛等平台动作。",
        "topic_signal": "《波斯复仇记》 / ToonScroll / PineDrama / TikTok minis / FlareFlow / AI漫剧工业化",
        "evidence_level": "B",
        "source_url": "https://www.36kr.com/p/3836864874560388",
    },
    {
        "source_name": "Business Insider",
        "article_date": "2026-05-13",
        "title": "Peacock is licensing micro dramas from ReelShort",
        "summary": "NBCUniversal 将 ReelShort 微短剧引入 Peacock 移动端，并规划 Bravo 原创 micro drama，可作为主流流媒体试水竖屏短剧的信号。",
        "topic_signal": "主流流媒体试水 / ReelShort授权 / Hollywood vertical drama",
        "evidence_level": "B",
        "source_url": "https://www.businessinsider.com/peacock-reelshort-micro-dramas-bravo-hollywood-short-form-vertical-video-2026-5",
    },
    {
        "source_name": "The Week",
        "article_date": "2026-02-18",
        "title": "Microdramas are booming",
        "summary": "文章从美国大众媒体视角解释 microdrama 的观看场景、付费模式、低成本制作和洛杉矶产业机会，可补充海外受众教育视角。",
        "topic_signal": "美国市场教育 / 竖屏付费短剧 / 低成本高周转",
        "evidence_level": "B",
        "source_url": "https://theweek.com/culture-life/film/microdramas-short-tiktok-entertainment",
    },
    {
        "source_name": "arXiv / HCI",
        "article_date": "2026-02-15",
        "title": "Audience in the Loop: Viewer Feedback-Driven Content Creation in Micro-drama Production on Social Media",
        "summary": "论文从编剧和生产流程角度讨论短剧创作如何吸收评论、转发、表情包等观众反馈，可作为题材迭代方法论观察。",
        "topic_signal": "观众反馈驱动创作 / 评论共创 / 短剧生产流程",
        "evidence_level": "C",
        "source_url": "https://arxiv.org/abs/2602.14045",
    },
    {
        "source_name": "Business Insider",
        "article_date": "2026-01-31",
        "title": "Creators and fans wrestle with the dark side of the micro drama boom",
        "summary": "文章关注 micro drama 中暴力、羞辱和厌女桥段的争议，以及粉丝对更强女性角色、多元选角和角色驱动故事的呼声。",
        "topic_signal": "题材风险 / 女性角色升级 / 反羞辱叙事",
        "evidence_level": "B",
        "source_url": "https://www.businessinsider.com/micro-dramas-spark-backlash-over-violence-and-misogyny-in-plots-2026-1",
    },
    {
        "source_name": "Axios",
        "article_date": "2026-01-15",
        "title": "Holywater raises $22M for microdramas",
        "summary": "Holywater 为 My Drama 等 microdrama 产品获得 2200 万美元 A 轮融资，说明非亚洲团队也在加码竖屏短剧平台。",
        "topic_signal": "海外平台融资 / My Drama / 非亚洲玩家入局",
        "evidence_level": "B",
        "source_url": "https://www.axios.com/2026/01/15/holywater-microdrama-app-funding",
    },
    {
        "source_name": "DataEye短剧观察",
        "article_date": "2026-06-01",
        "title": "DataEye海外微短剧热榜：“女版五十度灰”跻身TOP3，两部“女主逆袭”新剧上榜",
        "summary": "海外微短剧热榜结果显示情欲悬疑、女主逆袭等题材仍有高热度，适合追踪海外付费短剧用户的强情绪偏好。",
        "topic_signal": "海外女频：情欲悬疑 / 女主逆袭 / 高压关系",
        "evidence_level": "B",
        "source_url": "https://weixin.sogou.com/weixin?type=2&query=DataEye%20%E7%9F%AD%E5%89%A7",
    },
    {
        "source_name": "DataEye短剧观察",
        "article_date": "2026-03-30",
        "title": "DataEye短剧&漫剧日榜：6部九州发行AI仿真人剧居前列，红果漫剧榜TOP10有3部萌宝",
        "summary": "搜索结果显示AI仿真人剧、萌宝漫剧在短剧/漫剧日榜中集中出现，可用于观察AI形态与经典高转化题材的结合。",
        "topic_signal": "AI仿真人剧 / 萌宝漫剧 / 榜单集中度",
        "evidence_level": "C",
        "source_url": "https://weixin.sogou.com/weixin?type=2&query=DataEye%20%E7%9F%AD%E5%89%A7",
    },
    {
        "source_name": "DataEye短剧出海",
        "article_date": "2026-03-27",
        "title": "DataEye海外微短剧热榜：西幻AI新剧上榜，ReelShort登顶素材增长榜首",
        "summary": "海外短剧榜单出现西幻、AI相关新剧，同时素材增长榜能观察投放侧正在加码的类型。",
        "topic_signal": "西幻短剧 / AI新剧 / 投放素材增长",
        "evidence_level": "B",
        "source_url": "https://weixin.sogou.com/weixin?type=2&query=DataEye%20%E7%9F%AD%E5%89%A7",
    },
    {
        "source_name": "DataEye短剧出海",
        "article_date": "2026-03-15",
        "title": "DataEye海外微短剧热榜：“虐恋追妻”新剧上榜，麦芽、昆仑万维包揽双榜TOP3",
        "summary": "海外短剧热榜提到虐恋追妻、平台发行方榜单变化，可继续观察追妻火葬场和头部供给竞争。",
        "topic_signal": "虐恋追妻 / 头部平台供给 / 海外女频",
        "evidence_level": "B",
        "source_url": "https://weixin.sogou.com/weixin?type=2&query=DataEye%20%E7%9F%AD%E5%89%A7",
    },
    {
        "source_name": "DataEye短剧观察",
        "article_date": "2026-03-22",
        "title": "DataEye短剧&漫剧日榜：2D漫包揽TOP2，5部短剧播放增量超1亿，多剧同登3榜",
        "summary": "2D漫剧和短剧播放增量同时被榜单提及，可作为漫剧和短剧交叉题材的观察入口。",
        "topic_signal": "2D漫剧 / 播放增量破亿 / 多榜共振",
        "evidence_level": "B",
        "source_url": "https://weixin.sogou.com/weixin?type=2&query=DataEye%20%E7%9F%AD%E5%89%A7",
    },
    {
        "source_name": "DataEye短剧观察",
        "article_date": "2026-03-08",
        "title": "DataEye短剧&漫剧日榜：“诡异”题材连续3期登顶，TOP2单日播放破亿",
        "summary": "诡异题材连续登顶说明悬疑惊悚、怪谈和反常识设定值得作为短剧/漫剧方向观察。",
        "topic_signal": "诡异题材 / 怪谈悬疑 / 高播放增量",
        "evidence_level": "B",
        "source_url": "https://weixin.sogou.com/weixin?type=2&query=DataEye%20%E7%9F%AD%E5%89%A7",
    },
    {
        "source_name": "DataEye短剧观察",
        "article_date": "2026-03-01",
        "title": "DataEye短剧&漫剧日榜：AI仿真人剧包揽TOP5，《少夫人来自东北》空降榜首",
        "summary": "AI仿真人剧在日榜中集中占位，东北地域、少夫人、付免双爆等信号可用于观察AI短剧题材包装。",
        "topic_signal": "AI仿真人剧 / 地域爽剧 / 付免双爆",
        "evidence_level": "B",
        "source_url": "https://weixin.sogou.com/weixin?type=2&query=DataEye%20%E7%9F%AD%E5%89%A7",
    },
]


def build_ai_animation_topics():
    topics = [dict(item) for item in AI_ANIMATION_TOPIC_SEEDS]
    if topics:
        return topics

    if OUTPUT_PATH.exists():
        try:
            existing_payload = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
            return existing_payload.get("ai_animation_topics", [])[:6]
        except (OSError, json.JSONDecodeError):
            return []

    return []


def build_traditional_film_tv_topics():
    topics = [dict(item) for item in TRADITIONAL_FILM_TV_TOPIC_SEEDS]
    if topics:
        return topics

    if OUTPUT_PATH.exists():
        try:
            existing_payload = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
            return existing_payload.get("traditional_film_tv_topics", [])[:6]
        except (OSError, json.JSONDecodeError):
            return []

    return []


def build_industry_media_observations():
    def sort_key(item):
        is_dataeye = "DataEye" in str(item.get("source_name", ""))
        is_june_2026 = str(item.get("article_date", "")).startswith("2026-06")
        try:
            article_day = date.fromisoformat(str(item.get("article_date", ""))).toordinal()
        except ValueError:
            article_day = 0
        return (0 if is_june_2026 else 1, -article_day, 0 if is_dataeye else 1)

    observations = [
        dict(item)
        for item in INDUSTRY_MEDIA_OBSERVATION_SEEDS
        if str(item.get("article_date", "")).startswith("2026")
    ]
    observations.sort(key=sort_key)
    if observations:
        return observations

    if OUTPUT_PATH.exists():
        try:
            existing_payload = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
            existing_observations = [
                item
                for item in existing_payload.get("industry_media_observations", [])
                if str(item.get("article_date", "")).startswith("2026")
            ]
            existing_observations.sort(key=sort_key)
            return existing_observations[:10]
        except (OSError, json.JSONDecodeError):
            return []

    return []


def main():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    generated_at = datetime.now()
    snapshot_name = f"{generated_at.date().isoformat()}.json"
    clusters = extend_clusters(sheet_records(workbook, "topic_clusters", 4, 5))
    angles = sheet_records(workbook, "drama_angle_map", 4, 5)
    watchlist = extend_watchlist(
        hydrate_watchlist(sheet_records(workbook, "weekly_watchlist", 4, 5), clusters, angles)
    )
    youtube_trending_videos, youtube_fetch_status = fetch_youtube_trending_videos(generated_at)
    youtube_comment_pain_points, youtube_comment_status = fetch_youtube_comment_pain_points(generated_at)
    signals = build_signals(sheet_records(workbook, "raw_signals", 4, 5), generated_at)

    payload = {
        "generated_at": generated_at.isoformat(timespec="seconds"),
        "source_workbook": WORKBOOK_PATH.name,
        "scope": {
            "markets": ["US", "UK", "CA", "AU"],
            "language": "EN",
            "cadence": "Manual or scheduled refresh; trailing 1/7/30 days selectable",
            "data_boundary": "Public pages and aggregate metrics only",
        },
        "snapshot": {
            "date": generated_at.date().isoformat(),
            "path": f"data/snapshots/{snapshot_name}",
        },
        "watchlist": with_audience_segments(watchlist),
        "clusters": with_audience_segments(clusters),
        "angles": with_audience_segments(angles),
        "signals": with_audience_segments(signals),
        "youtube_trending_videos": youtube_trending_videos,
        "youtube_fetch_status": youtube_fetch_status,
        "youtube_comment_status": youtube_comment_status,
        "ai_animation_topics": with_audience_segments(build_ai_animation_topics()),
        "traditional_film_tv_topics": with_audience_segments(build_traditional_film_tv_topics()),
        "comment_pain_points": with_audience_segments(build_comment_pain_points(youtube_comment_pain_points)),
        "industry_media_observations": with_audience_segments(build_industry_media_observations()),
        "strategic_focus": with_audience_segments(build_strategic_focus()),
        "dictionary": sheet_records(workbook, "dictionary", 4, 5),
        "weights": [
            {"name": "热度", "weight": 35},
            {"name": "情绪强度", "weight": 25},
            {"name": "题材可短剧化", "weight": 20},
            {"name": "供给缺口", "weight": 10},
            {"name": "合规/品牌风险扣分", "weight": 10},
        ],
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (SNAPSHOT_DIR / snapshot_name).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    snapshots = []
    for path in sorted(SNAPSHOT_DIR.glob("*.json"), reverse=True):
        if path.name == "index.json":
            continue
        snapshots.append({"date": path.stem, "path": f"data/snapshots/{path.name}"})
    SNAPSHOT_INDEX_PATH.write_text(json.dumps(snapshots, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {OUTPUT_PATH}")
    print(f"Wrote {SNAPSHOT_DIR / snapshot_name}")
    print(f"Wrote {SNAPSHOT_INDEX_PATH}")


if __name__ == "__main__":
    main()
